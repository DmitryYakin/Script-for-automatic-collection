import time
import schedule
from datetime import datetime
import random
import pandas as pd
import logging
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
import requests  # Убедитесь, что библиотека установлена (pip install requests)

# Конфигурация Telegram (замените эти значения)
TELEGRAM_BOT_TOKEN = "7969716964:AAGqiyxi6iEe9E5uxyDWV-zQLAYhLqx0i9E"  # Получить у @BotFather
TELEGRAM_CHAT_ID = "293265468"  # Узнать через @userinfobot или getUpdates




BASE_URL = "https://cloud.webiomed.ru/#/dhra/requests/"
EXCEL_FILE = "Статистика время открытия СППВР.xlsx"
SHEET_NAME = "02.06-08.06"





# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def login(browser):
    try:
        creds = pd.read_csv('creds.csv').loc[0]
        time.sleep(10)
        WebDriverWait(browser, 240).until(EC.presence_of_element_located((By.ID, "login")))
        browser.find_element(By.ID, "login").send_keys(creds['username'])

        WebDriverWait(browser, 30).until(EC.presence_of_element_located((By.ID, "password")))
        browser.find_element(By.ID, "password").send_keys(creds['password'] + Keys.RETURN)

        # Ждём переход после логина
        WebDriverWait(browser, 240).until(
            EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "menu")]'))
        )
        logging.info("Успешный вход в систему.")
    except Exception as e:
        logging.error(f"Ошибка при авторизации: {e}")
        browser.save_screenshot("error_login.png")
        raise

def write_to_excel(current_time, load_time, assessment_id, comment=""):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb[SHEET_NAME]

        today_str = datetime.now().strftime("%d.%m.%Y")
        row_found = False

        for row in range(4, sheet.max_row + 2):
            cell_date = sheet.cell(row=row, column=2).value  # колонка B

            if isinstance(cell_date, datetime):
                cell_date_str = cell_date.strftime("%d.%m.%Y")
            elif isinstance(cell_date, str):
                cell_date_str = cell_date.strip()
            else:
                continue

            if cell_date_str == today_str:
                # Если колонка C (время) пуста — используем эту строку
                if not sheet.cell(row=row, column=3).value:
                    sheet.cell(row=row, column=3).value = current_time  # Время (C)
                    sheet.cell(row=row, column=4).value = load_time if load_time is not None else ""  # Время открытия (D)
                    sheet.cell(row=row, column=7).value = assessment_id  # ID оценки (G)
                    sheet.cell(row=row, column=8).value = comment  # Комментарий (H)
                    row_found = True
                    break

        if row_found:
            wb.save(EXCEL_FILE)
            logging.info(f"✅ Запись добавлена: {today_str}, {current_time}, {load_time}, {assessment_id}, {comment}")
        else:
            logging.warning(f"⚠ Не найдена пустая строка с сегодняшней датой {today_str}")

    except Exception as e:
        logging.error(f"❌ Ошибка записи в Excel: {e}")

def send_telegram_alert(measure_time, assessment_id, load_time):
    """Отправляет персональное сообщение через Telegram-бота"""
    message = (
        f"В {measure_time} оценка {assessment_id} "
        f"открылась за {load_time} секунд"
    )

    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        params = {
            'chat_id': TELEGRAM_CHAT_ID,
            'text': message,
            'parse_mode': 'HTML'
        }

        response = requests.post(url, json=params)
        if response.status_code != 200:
            logging.error(f"Telegram API error: {response.text}")
    except Exception as e:
        logging.error(f"Ошибка отправки в Telegram: {e}")


def measure_load_time():
    driver = None
    now = datetime.now()
    current_time_str = now.strftime("%H:%M")
    assessment_id = random.randint(550000000, 600000000)
    url = f"{BASE_URL}{assessment_id}"

    try:
        options = webdriver.FirefoxOptions()
        options.add_argument("--disable-gpu")
        options.add_argument("--headless")
        driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()), options=options)

        driver.get(url)
        login(driver)

        start_time = time.time()
        WebDriverWait(driver, 240).until(EC.all_of(
EC.presence_of_element_located((By.XPATH,'//span[contains(@class,"section-header__title") and contains(text(),"Основные данные")]')),
                 EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div/div/div/div/div/div/div/div[1]/div/div[2]/div[2]/div/div[2]/div/div[1]')),                 # Cholesterol
                EC.presence_of_element_located((By.XPATH, '/html/body/div/div[1]/div/div/div/div/div/div/div/div[1]/div/div[2]/div[1]/div/div[2]/div'))  ))

        load_time = round(time.time() - start_time, 2)
        write_to_excel(current_time_str, load_time, assessment_id)

        # Отправляем уведомление в Telegram
        send_telegram_alert(current_time_str, assessment_id, load_time)

        driver.quit()
    except Exception as e:
        logging.error(f"Ошибка при измерении: {e}")
        if driver:
            driver.save_screenshot("error_load.png")
            driver.quit()
        write_to_excel(current_time_str, None, assessment_id, "Ошибка открытия")
        send_telegram_alert(current_time_str, assessment_id, "не открылась")

# Настройка расписания с 9:00 до 18:00 по минутам
for hour in range(9, 19):
    for minute in range(0, 3):
        schedule.every().day.at(f"{hour:02d}:{minute:02d}").do(measure_load_time)

logging.info("Скрипт запущен. Ожидание времени для проверки...")
while True:
    schedule.run_pending()
    time.sleep(1)
